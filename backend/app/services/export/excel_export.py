"""Exportación a Excel. El pipeline actual nunca genera contactos 'inferred'
(solo direct_published/company_general/not_found), así que no hace falta un
filtro adicional para no exportarlos como verificados -- si en el futuro se
agrega inferencia de patrones de correo, este exportador deberá excluir
explícitamente cualquier contact_type == 'inferred'.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.config import settings

STATUS_ES = {
    "candidate": "Candidato",
    "probable": "Probable",
    "confirmed": "Confirmado",
    "outdated": "Fuente antigua",
    "needs_review": "Requiere revisión",
    "discarded": "Descartado",
}

COLUMNS = [
    ("full_name", "Nombre"),
    ("job_title", "Cargo"),
    ("area", "Área"),
    ("company_name", "Empresa"),
    ("industry", "Sector"),
    ("country", "País"),
    ("city", "Ciudad"),
    ("direct_email", "Correo directo"),
    ("direct_email_type", "Tipo de correo"),
    ("company_email", "Correo corporativo"),
    ("direct_phone", "Teléfono directo"),
    ("direct_phone_type", "Tipo de teléfono"),
    ("company_phone", "Teléfono de empresa"),
    ("linkedin_url", "LinkedIn"),
    ("source_url", "Fuente"),
    ("source_date", "Fecha de la fuente"),
    ("evidence_text", "Evidencia"),
]

BRAND_FILL = PatternFill(start_color="FF003558", end_color="FF003558", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
MAX_COLUMN_WIDTH = 60


def build_export_dataframe(results: list[dict]) -> pd.DataFrame:
    rows = []
    for result in results:
        data = result["result_data_json"]
        row = {label: data.get(key) for key, label in COLUMNS}
        row["Estado"] = STATUS_ES.get(result["status"], result["status"])
        row["Campos faltantes"] = ", ".join(data.get("missing_fields") or [])
        rows.append(row)
    return pd.DataFrame(rows)


def _style_as_table(worksheet, df: pd.DataFrame, table_name: str) -> None:
    """Da formato de tabla real de Excel: encabezado azul Tecport, franjas
    alternadas, autofiltro y columnas con ancho ajustado al contenido."""
    n_rows, n_cols = df.shape
    last_col_letter = get_column_letter(n_cols)

    for col_idx, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = BRAND_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

        max_len = max(
            [len(str(column))] + [len(str(value)) for value in df[column].fillna("") if str(value)] or [0]
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COLUMN_WIDTH)

    worksheet.freeze_panes = "A2"

    if n_rows > 0:
        table_ref = f"A1:{last_col_letter}{n_rows + 1}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def export_search_to_excel(search_id: int, results: list[dict], file_name: str) -> str:
    export_dir = Path(settings.export_directory)
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / file_name

    df = build_export_dataframe(results)
    summary = (
        df["Estado"].value_counts().rename_axis("Estado").reset_index(name="Cantidad")
        if not df.empty
        else pd.DataFrame(columns=["Estado", "Cantidad"])
    )

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultados", index=False)
        summary.to_excel(writer, sheet_name="Resumen", index=False)

        _style_as_table(writer.sheets["Resultados"], df, f"Resultados_{search_id}")
        _style_as_table(writer.sheets["Resumen"], summary, f"Resumen_{search_id}")

    return str(file_path)
